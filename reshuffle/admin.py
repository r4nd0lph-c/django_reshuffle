from django.contrib import admin
from django.contrib.admin.helpers import ACTION_CHECKBOX_NAME
from django.http import HttpResponse
from django.utils import timezone
from django.utils.safestring import mark_safe

from .views import *
from .models import *
from .widgets import LatexInput

from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# Register your models here.


class SubjectsAdmin(admin.ModelAdmin):
    list_display = ('id', 'case_nominative', 'tasks_number')
    list_display_links = ('id', 'case_nominative')


class TasksForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super(TasksForm, self).__init__(*args, **kwargs)
        groups = self.request.user.groups.all()
        if not groups.exists() or self.request.user.is_superuser:
            pass
        else:
            self.fields['subject_fk'].queryset = Subjects.objects.filter(pk__in=subj_access_list(groups))

    def clean(self):
        if 'num' in self.cleaned_data:
            num = self.cleaned_data['num']
            if 'subject_fk' in self.cleaned_data:
                subj_id = self.cleaned_data['subject_fk'].id
                t_num = Subjects.objects.get(id=subj_id).tasks_number
                if num <= t_num:
                    if self.cleaned_data['text'] == '' and self.cleaned_data['latex'] == '' \
                            and (self.cleaned_data['image'] is None or self.cleaned_data['image'] is False):
                        raise forms.ValidationError({
                            'text': 'Хотя бы одно из полей "{}", "{}", "{}" должно быть заполнено.'.format(
                                Tasks._meta.get_field('text').verbose_name,
                                Tasks._meta.get_field('latex').verbose_name,
                                Tasks._meta.get_field('image').verbose_name)})
                else:
                    raise forms.ValidationError(
                        {'num': 'Убедитесь, что это значение меньше либо равно {}.'.format(t_num)})
            else:
                raise forms.ValidationError(
                    {'num': 'Для начала укажите "{}".'.format(Tasks._meta.get_field('subject_fk').verbose_name)})
        else:
            raise forms.ValidationError({'num': 'Убедитесь, что вы заполнили это поле.'})


class TasksAdmin(admin.ModelAdmin):
    form = TasksForm

    list_display = ('id', 'subject_fk', 'num', 'text', 'latex', 'get_html_img')
    list_display_links = ('id',)
    list_filter = (('subject_fk', admin.RelatedOnlyFieldListFilter),)
    search_fields = ('=num',)
    search_help_text = 'Поиск выполняется по номеру задания в тесте \
                        (для более точного результата вы можете использовать фильтры)'

    formfield_overrides = {
        LatexField: {'widget': LatexInput},
    }

    def get_queryset(self, request):
        qs = super(TasksAdmin, self).get_queryset(request)
        groups = request.user.groups.all()
        if not groups.exists() or request.user.is_superuser:
            return qs
        else:
            return qs.filter(subject_fk__id__in=subj_access_list(groups))

    def get_form(self, request, *args, **kwargs):
        form = super(TasksAdmin, self).get_form(request, *args, **kwargs)
        form.request = request
        return form

    def get_html_img(self, object):
        if object.image:
            return mark_safe('<img src="{}" style="max-width:100%">'.format(object.image.url))

    get_html_img.short_description = 'Изображение'


class OptionsForm(forms.ModelForm):
    def clean(self):
        if self.cleaned_data['text'] == '' and self.cleaned_data['latex'] == '' \
                and (self.cleaned_data['image'] is None or self.cleaned_data['image'] is False):
            raise forms.ValidationError({
                'text': 'Хотя бы одно из полей "{}", "{}", "{}" должно быть заполнено.'.format(
                    Tasks._meta.get_field('text').verbose_name,
                    Tasks._meta.get_field('latex').verbose_name,
                    Tasks._meta.get_field('image').verbose_name)})


class OptionsAdmin(admin.ModelAdmin):
    form = OptionsForm

    list_display = ('id', 'task_fk', 'text', 'latex', 'get_html_img', 'is_answer')
    list_display_links = ('id',)
    list_filter = (('task_fk__subject_fk', admin.RelatedOnlyFieldListFilter), 'is_answer')
    search_fields = ('=task_fk__id',)
    search_help_text = 'Поиск выполняется по id задания, к которому относятся ответы'
    raw_id_fields = ('task_fk',)

    # autocomplete_fields = ['task_fk']

    formfield_overrides = {
        LatexField: {'widget': LatexInput},
    }

    def get_queryset(self, request):
        qs = super(OptionsAdmin, self).get_queryset(request)
        groups = request.user.groups.all()
        if not groups.exists() or request.user.is_superuser:
            return qs
        else:
            return qs.filter(task_fk__subject_fk__id__in=subj_access_list(groups))

    def get_html_img(self, object):
        if object.image:
            return mark_safe('<img src="{}" style="max-width:100%">'.format(object.image.url))

    get_html_img.short_description = 'Изображение'


def subj_access_list(groups):
    access_list = []
    for group in groups:
        qs = SubjAccess.objects.filter(group_id=group.id)
        for item in qs:
            access_list.append(item.subject_fk_id)
    return access_list


class SubjAccessAdmin(admin.ModelAdmin):
    list_display = ('group_id', 'subject_fk')
    list_display_links = ('group_id',)


class ArchiveLogsDateFilter(admin.SimpleListFilter):
    title = 'Прошедшее время'
    parameter_name = 'action_date'

    def lookups(self, request, model_admin):
        return (
            ('last_day', 'Меньше суток'),
            ('last_three_days', 'Меньше трёх дней'),
            ('last_week', 'Меньше недели'),
        )

    def queryset(self, request, queryset):
        timestamps = {
            'last_day': 1,
            'last_three_days': 3,
            'last_week': 7
        }
        if self.value() is not None:
            start_date = timezone.now() - timedelta(days=timestamps[self.value()])
            return queryset.filter(action_time__gte=start_date)
        return queryset


class ArchiveLogsAdmin(admin.ModelAdmin):
    list_display = ('action_time', 'username', 'archive_info', 'action')
    list_display_links = None
    list_filter = (('username', admin.RelatedOnlyFieldListFilter), ArchiveLogsDateFilter)
    actions = ['download_logs']

    def has_add_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.action(description='Скачать элементы (если выбрано 0, то скачается весь журнал) в формате XLSX')
    def download_logs(self, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', )
        response['Content-Disposition'] = 'attachment; filename=logs_{date}.xlsx'.format(
            date=datetime.now().strftime('%Y_%m_%d_%H_%M_%S_%f'))

        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Журнал действий'

        # Define the titles for columns
        columns = [
            ArchiveLogs._meta.get_field('action_time').verbose_name.upper(),
            ArchiveLogs._meta.get_field('username').verbose_name.upper(),
            ArchiveLogs._meta.get_field('archive_info').verbose_name.upper(),
            ArchiveLogs._meta.get_field('action').verbose_name.upper()
        ]

        for i in range(len(columns)):
            column_letter = get_column_letter(i + 1)
            column_dimensions = worksheet.column_dimensions[column_letter]
            column_dimensions.width = 32 if i == 2 else 16

        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(bold=True)

        # Iterate through all movies
        for item in queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                item.action_time.strftime('%d.%m.%Y %H:%M'),
                item.username.username,
                item.archive_info,
                item.action
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response

    def changelist_view(self, request, extra_context=None):
        if 'action' in request.POST and request.POST['action'] == 'download_logs':
            if not request.POST.getlist(ACTION_CHECKBOX_NAME):
                post = request.POST.copy()
                for u in ArchiveLogs.objects.all():
                    post.update({ACTION_CHECKBOX_NAME: str(u.id)})
                request._set_post(post)
        return super(ArchiveLogsAdmin, self).changelist_view(request, extra_context)


admin.site.register(Subjects, SubjectsAdmin)
admin.site.register(Tasks, TasksAdmin)
admin.site.register(Options, OptionsAdmin)
admin.site.register(SubjAccess, SubjAccessAdmin)
admin.site.register(ArchiveLogs, ArchiveLogsAdmin)

admin.site.site_title = 'RESHUFFLE'
admin.site.site_header = 'RESHUFFLE'
admin.site.index_title = 'Администрирование приложения'

admin.site.login = Auth.as_view()
admin.site.logout = logout_user
